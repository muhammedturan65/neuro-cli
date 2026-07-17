#!/usr/bin/env python3
"""NeuroCLI vs Competitors Feature Comparison Matrix - XLSX"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== SHEET 1: Feature Comparison Matrix =====
ws = wb.active
ws.title = "Özellik Kıyaslaması"

# Colors
HDR_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HDR2_FILL = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
NEURO_FILL = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
YES_FILL = PatternFill(start_color="1b4332", end_color="1b4332", fill_type="solid")
NO_FILL = PatternFill(start_color="641220", end_color="641220", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="7f4f24", end_color="7f4f24", fill_type="solid")
ALT_ROW = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=False, size=10)
BOLD_WHITE = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="E94560", bold=True, size=14)
ACCENT_FONT = Font(color="00D2FF", bold=True, size=10)
GREEN_FONT = Font(color="52B788", bold=True, size=10)
RED_FONT = Font(color="E5383B", bold=True, size=10)
ORANGE_FONT = Font(color="E9C46A", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='333355'),
    right=Side(style='thin', color='333355'),
    top=Side(style='thin', color='333355'),
    bottom=Side(style='thin', color='333355')
)

# Title
ws.merge_cells('A1:G1')
ws['A1'] = 'NEUROCLI vs RAKIP ARACLAR - OZELLIK KIYASLAMA MATRISI'
ws['A1'].font = TITLE_FONT
ws['A1'].alignment = Alignment(horizontal='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:G2')
ws['A2'] = 'Temmuz 2026 | Detayli Arastirma Raporu'
ws['A2'].font = Font(color="888888", size=9)
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Kategori', 'Ozellik', 'NeuroCLI', 'Claude Code', 'Gemini CLI', 'OpenCode/Crush', 'Kilo Code']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = BOLD_WHITE
    c.fill = HDR_FILL if col < 3 else (NEURO_FILL if col == 3 else HDR2_FILL)
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = THIN_BORDER

# Column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 32
for col_letter in ['C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col_letter].width = 22

# Data rows
features = [
    # Category, Feature, NeuroCLI, Claude Code, Gemini CLI, OpenCode, Kilo Code
    # --- TEMEL OZELLIKLER ---
    ("Temel Ozellikler", "Interaktif REPL Modu", "✅ Tam", "✅ Tam", "✅ Tam", "✅ Tam (TUI)", "✅ Tam"),
    ("Temel Ozellikler", "Tek Seferlik (One-Shot) Mod", "✅ `neuro ask`", "✅ `-p` flag", "✅ `gemini -p`", "✅ `opencode -p`", "✅ `kilo run`"),
    ("Temel Ozellikler", "Coklu Dosya Duzenleme", "✅ Tam", "✅ Tam", "✅ Tam", "✅ Tam", "✅ Tam"),
    ("Temel Ozellikler", "Otonom Kod Uretimi", "✅ Tam", "✅ Tam", "✅ Tam", "✅ Tam", "✅ Tam"),
    ("Temel Ozellikler", "Headless/CI Modu", "❌ Yok", "✅ `--max-turns`", "✅ Headless", "✅ `-p -f json`", "✅ `--auto`"),
    ("Temel Ozellikler", "Arka Plan Agentleri", "❌ Yok", "✅ Background agents", "❌ Yok", "❌ Yok", "❌ Yok"),
    
    # --- ARAC MIMARISI ---
    ("Arac Mimarisi", "Yerlesik Ajan Sayisi", "8", "1 + sub-agents", "4 + custom", "4 + custom", "5 + custom"),
    ("Arac Mimarisi", "Sub-agent Destegi", "✅ Team+Orchestrator", "✅ Agent tool", "✅ Hub-Spoke", "✅ @mention", "✅ task tool"),
    ("Arac Mimarisi", "Paralel Calistirma", "✅ DAG tabanli", "✅ Paralel subagents", "✅ Paralel", "✅ General subagent", "✅ Agent Manager"),
    ("Arac Mimarisi", "Ozel Ajan Tanimlama", "⚠️ Config tabanli", "✅ YAML frontmatter", "✅ Agent def files", "✅ .opencode/agents/", "✅ kilo agent create"),
    ("Arac Mimarisi", "Ajanlar Arasi Mesajlasma", "✅ 5 mesaj tipi", "❌ Sonuc raporu", "❌ Sonuc raporu", "❌ Sonuc raporu", "❌ Sonuc raporu"),
    ("Arac Mimarisi", "Uzaktan Ajan (A2A)", "❌ Yok", "❌ Yok", "✅ A2A protokol", "❌ Yok", "⚠️ ACP server"),
    
    # --- ARAK KAPASITELERI ---
    ("Arac Kapasiteleri", "Dosya Okuma/Yazma", "✅ 7 tool", "✅ Read/Write/Edit", "✅ Read/Write", "✅ read/write/edit", "✅ read/edit"),
    ("Arac Kapasiteleri", "Bash/Kabuk Calistirma", "✅ run_command", "✅ Bash tool", "✅ Shell tool", "✅ bash tool", "✅ bash tool"),
    ("Arac Kapasiteleri", "Git Entegrasyonu", "✅ 18 alt-komut", "✅ Derin git", "✅ Shell uzerinden", "✅ bash uzerinden", "✅ PR checkout dahil"),
    ("Arac Kapasiteleri", "Web Arama", "✅ DuckDuckGo", "✅ Web Search", "✅ Google Search", "✅ websearch", "✅ webfetch"),
    ("Arac Kapasiteleri", "Web Sayfa Cekme", "✅ web_fetch", "✅ Web Reader", "✅ Web Fetch", "✅ webfetch", "✅ webfetch"),
    ("Arac Kapasiteleri", "Kod Arama (Grep)", "✅ ripgrep", "✅ Grep tool", "✅ ripgrep", "✅ grep tool", "✅ grep tool"),
    ("Arac Kapasiteleri", "Diff/Uygulama", "✅ apply_diff", "✅ Multi-edit", "⚠️ Edit tool", "✅ apply_patch", "✅ edit tool"),
    ("Arac Kapasiteleri", "LSP Entegrasyonu", "✅ 4 dil (TS/Py/Go/Rust)", "❌ Yok", "❌ Yok", "✅ Deneysel LSP", "❌ Yok"),
    ("Arac Kapasiteleri", "Tarayici Ajanı", "❌ Yok", "✅ Chrome ent.", "✅ Browser Agent", "❌ Yok", "⚠️ MCP uzerinden"),
    
    # --- BAGLAM YONETIMI ---
    ("Baglam Yonetimi", "Baglam Penceresi", "Model bagimli", "200K-1M", "1M", "Model bagimli", "Model bagimli"),
    ("Baglam Yonetimi", "Otomatik Sikistirma", "✅ 5 katmanli", "✅ Auto-compaction", "⚠️ Subagent izolasyonu", "✅ Auto Compact 95%", "✅ Auto-Compaction"),
    ("Baglam Yonetimi", "Manuel Sikistirma", "✅ /compact", "✅ /compact", "❌ Yok", "✅ Leader+c", "✅ /compact"),
    ("Baglam Yonetimi", "Proje Baglam Dosyasi", "✅ NEURO.md (5 katman)", "✅ CLAUDE.md", "✅ GEMINI.md", "✅ AGENTS.md", "✅ AGENTS.md"),
    ("Baglam Yonetimi", "Dosya Izleme/Canli Guncelleme", "✅ 5s polling", "❌ Yok", "❌ Yok", "❌ Yok", "❌ Yok"),
    ("Baglam Yonetimi", "Repo Haritasi", "✅ 5 dil destegi", "❌ Yok", "❌ Yok", "❌ Yok", "✅ Codebase Indexing"),
    ("Baglam Yonetimi", "Bellek/Kalici Hafiza", "✅ save/recall_memory", "❌ Yok", "✅ Auto Memory", "❌ Yok", "⚠️ Deprecated Memory Bank"),
    
    # --- MCP DESTEKI ---
    ("MCP Destegi", "MCP Protokol Destegi", "❌ YOK", "✅ 3000+ entegrasyon", "✅ SSE/HTTP/stdio", "✅ Local+Remote+OAuth", "✅ Full + Marketplace"),
    ("MCP Destegi", "MCP Yapilandirma", "❌ YOK", "✅ .mcp.json + CLI", "✅ settings.json", "✅ opencode.json", "✅ kilo.jsonc"),
    ("MCP Destegi", "MCP Marketplace", "❌ YOK", "❌ Yok", "❌ Yok", "❌ Yok", "✅ kilo-marketplace"),
    
    # --- IZIN VE GUVENLIK ---
    ("Izin ve Guvenlik", "Onay Modlari", "⚠️ 2 mod (auto/manual)", "✅ 3 mod", "✅ 4 mod", "✅ 3 kademe", "✅ Granular per-tool"),
    ("Izin ve Guvenlik", "Tehlikeli Komut Algilama", "✅ Regex tabanli", "✅ AI siniflandirici", "⚠️ Workspace trust", "⚠️ Pattern tabanli", "✅ Sandbox + pattern"),
    ("Izin ve Guvenlik", "Sandbox/Yalitim", "❌ Yok", "✅ Seatbelt+Bubblewrap", "✅ 5 yontem", "❌ Yok", "✅ Seatbelt+Bubblewrap"),
    ("Izin ve Guvenlik", "Doom Loop Korumasi", "❌ Yok", "❌ Yok", "❌ Yok", "❌ Yok", "✅ Varsayilan"),
    
    # --- OTURUM YONETIMI ---
    ("Oturum Yonetimi", "Oturum Kaydetme", "✅ JSON dosya", "✅ Otomatik", "✅ Otomatik", "✅ SQLite", "✅ JSON"),
    ("Oturum Yonetimi", "Oturum Devam Ettirme", "⚠️ load() var, CLI yok", "✅ -c / -r", "✅ --resume", "✅ Ctrl+A", "✅ --continue"),
    ("Oturum Yonetimi", "Oturum Fork", "❌ Yok", "✅ --fork-session", "✅ /resume save", "✅ Fork support", "✅ --fork"),
    ("Oturum Yonetimi", "Oturum Paylasma", "❌ Yok", "✅ /export", "❌ Yok", "✅ /share", "✅ /share + URL"),
    ("Oturum Yonetimi", "Oturum Tarayicisi", "❌ Yok", "✅ /resume picker", "✅ Interactive", "✅ Session list", "✅ /sessions"),
    
    # --- IDE ENTEGRASYONU ---
    ("IDE Entegrasyonu", "VS Code Eklentisi", "❌ YOK", "✅ Native extension", "✅ Companion ext", "✅ Official ext", "✅ 1M+ install"),
    ("IDE Entegrasyonu", "JetBrains Eklentisi", "❌ YOK", "✅ Dedicated plugin", "✅ ACP protokol", "❌ Terminal only", "✅ Kotlin plugin"),
    ("IDE Entegrasyonu", "Diff Goruntuleyici", "❌ YOK", "✅ Native diff viewer", "✅ VS Code diff", "✅ TUI diff wrap", "⚠️ v7 regression"),
    ("IDE Entegrasyonu", "Masaustu Uygulama", "❌ YOK", "✅ Desktop app", "❌ Yok", "✅ Desktop + Web", "✅ Cloud Agent"),
    
    # --- UZANTI SISTEMI ---
    ("Uzanti Sistemi", "Ozel Slash Komutlari", "✅ 14+ komut", "✅ .claude/commands/", "✅ Custom commands", "✅ .opencode/commands/", "✅ Workflows"),
    ("Uzanti Sistemi", "Hooks Sistemi", "✅ 20 olay, 4 tip", "✅ Pre/Post hooks", "✅ Lifecycle hooks", "✅ Before/After", "⚠️ Basit"),
    ("Uzanti Sistemi", "Ozel Araclar", "❌ Yok", "✅ Plugin SDK", "✅ Extensions", "✅ @opencode-ai/plugin", "✅ Custom tools"),
    ("Uzanti Sistemi", "Beceri (Skill) Sistemi", "❌ Yok", "✅ Auto-activate skills", "✅ Agent Skills", "✅ .opencode/skills/", "✅ .kilo/skills/"),
    ("Uzanti Sistemi", "Eklenti/Pazaryeri", "❌ Yok", "✅ 28+ plugin", "✅ Extension registry", "❌ Yok", "✅ Kilo Marketplace"),
    
    # --- UI/UX ---
    ("UI/UX", "Tema Sayisi", "4 (Dracula/Dark/Nord/Light)", "2 (Dark/Light)", "3 (Dark/Light/Auto)", "11+ tema + ozel", "Yapilandirilabilir"),
    ("UI/UX", "Sozdizimi Vurgulama", "❌ Yok", "✅ Ctrl+T toggle", "✅ Tema tabanli", "✅ TUI tabanli", "✅ Chat icinde"),
    ("UI/UX", "Akici Cikti (Streaming)", "✅ SSE", "✅ Token-by-token", "✅ Streaming", "✅ Streaming", "✅ Streaming"),
    ("UI/UX", "Otomatik Tamamlama", "❌ YOK", "✅ Slash+shell+model", "✅ Slash+@", "✅ Slash+@", "✅ FIM autocomplete"),
    ("UI/UX", "Sesli Giris/Cikti", "❌ Yok", "✅ Voice mode", "❌ Yok", "❌ Yok", "✅ Voice transcription"),
    ("UI/UX", "Ilk Durum Cubugu", "❌ Yok", "✅ /statusline", "❌ Yok", "✅ Leader+s", "❌ Yok"),
    
    # --- MODEL DESTEGI ---
    ("Model Destegi", "Ucretsiz Model Sayisi", "23", "0 (ucretli)", "1 (sinirli gunluk)", "0 (BYOK)", "200 req/saat"),
    ("Model Destegi", "Toplam Model Sayisi", "36", "~5 (Anthropic)", "~8 (Gemini)", "75+ provider", "500+ (Gateway)"),
    ("Model Destegi", "Coklu Saglayici", "⚠️ Sadece OpenRouter", "❌ Sadece Anthropic", "❌ Sadece Google", "✅ 75+ provider", "✅ BYOK + Gateway"),
    ("Model Destegi", "Yerel Model (Ollama)", "❌ Yok", "❌ Yok", "❌ Yok", "✅ Ollama/LM Studio", "✅ Ollama/LM Studio"),
    ("Model Destegi", "Model Yonlendirme", "❌ Yok", "✅ Effort levels", "✅ Auto mode + Gemma router", "⚠️ Manuel secim", "✅ Auto model"),
    ("Model Destegi", "Gorev Sirasinda Model Degistirme", "❌ Yok", "✅ /model", "✅ /model", "✅ Ctrl+T", "✅ /models"),
    ("Model Destegi", "Fiyatlandirma (Gelistirici/Gun)", "$0 (ucretsiz modeller)", "$4-13/gun", "Ucretsiz (sinirli)", "Kullanim bazli", "$0.00+ BYOK"),
    
    # --- TOKEN/MALIYET ---
    ("Token/Maliyet", "Kullanim Izleme", "✅ Session bazli", "✅ /cost detayli", "✅ /stats", "⚠️ Basit", "✅ kilo stats"),
    ("Token/Maliyet", "Maliyet Tahmini", "✅ Per-model fiyat", "✅ $/session breakdown", "⚠️ Token sayisi", "❌ Yok", "✅ Per-request"),
    ("Token/Maliyet", "Harcama Limiti", "❌ Yok", "✅ Org limits", "✅ Gunluk limit", "❌ Yok", "✅ Per-user daily"),
    ("Token/Maliyet", "Onbellek (Cache)", "❌ Yok", "✅ Prompt caching", "✅ Token caching", "❌ Yok", "❌ Yok"),
    
    # --- GERI ALMA / DIFF ---
    ("Geri Alma / Diff", "Geri Alma (Undo)", "⚠️ Git checkpoint var, CLI yok", "✅ /rewind (Escx2)", "✅ /rewind + /restore", "✅ /undo + /redo", "✅ /undo + /redo"),
    ("Geri Alma / Diff", "Diff Onizleme", "❌ YOK", "✅ Inline + IDE diff", "✅ IDE diff viewer", "✅ TUI diff", "⚠️ Kismi"),
    ("Geri Alma / Diff", "Git Tabanli Geri Donus", "✅ Auto-checkpoint", "✅ Diffs before commit", "✅ Git worktrees", "✅ Git integration", "✅ Checkpoint revert"),
    
    # --- PROJE FARKINDALIGI ---
    ("Proje Farkindaligi", "Proje Baglam Dosyasi", "✅ NEURO.md", "✅ CLAUDE.md", "✅ GEMINI.md", "✅ AGENTS.md", "✅ AGENTS.md"),
    ("Proje Farkindaligi", "Hiyerarsik Baglam", "✅ 5 katman + @import", "✅ 4 katman", "✅ 3+ katman", "✅ Global + project", "✅ Global + per-dir"),
    ("Proje Farkindaligi", ".ignore Dosyasi", "❌ Yok", "❌ Yok", "✅ .geminiignore", "❌ Yok", "✅ .kilocodeignore"),
    ("Proje Farkindaligi", "/init Komutu", "✅ Var (auto-generate)", "❌ Yok", "❌ Yok", "✅ Repo tara+olustur", "❌ Yok"),
    ("Proje Farkindaligi", "Otomatik Teknoloji Tespiti", "✅ project_context", "❌ Yok", "❌ Yok", "❌ Yok", "⚠️ Codebase Indexing"),
    
    # --- DANISMAN SISTEMI ---
    ("Danisman Sistemi", "Ikinci Model Danismani", "✅ 6 tetikleyici", "✅ Advisor tool", "❌ Yok", "❌ Yok", "❌ Yok"),
    ("Danisman Sistemi", "Tekrar Eden Hata Algilama", "✅ Kelime benzerligi", "❌ Yok", "❌ Yok", "❌ Yok", "❌ Yok"),
    
    # --- GENEL PUAN ---
    ("GENEL DEGERLENDIRME", "Toplam Ozellik Sayisi (yaklasik)", "~45", "~65", "~55", "~60", "~70"),
    ("GENEL DEGERLENDIRME", "Benzersiz Avantaj", "Ucretsiz modeller + Danisman + LSP + 5 katmanli sikistirma", "En olgun ekosistem + 3000+ MCP", "1M baglam + A2A + 5 sandbox", "Go TUI + 75+ provider + Ozel arac SDK", "Sandbox + Marketplace + 500+ model"),
]

row = 5
current_cat = ""
for cat, feat, neuro, claude, gemini, opencode, kilo in features:
    # Category grouping
    if cat != current_cat:
        current_cat = cat
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=cat)
        c.font = Font(color="00D2FF", bold=True, size=11)
        c.fill = PatternFill(start_color="0a0a1a", end_color="0a0a1a", fill_type="solid")
        c.alignment = Alignment(horizontal='center')
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # Feature name
    ws.cell(row=row, column=1, value="").border = THIN_BORDER
    ws.cell(row=row, column=2, value=feat).font = WHITE_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    
    # Tool values with color coding
    for col, val in [(3, neuro), (4, claude), (5, gemini), (6, opencode), (7, kilo)]:
        c = ws.cell(row=row, column=col, value=val)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        
        if val.startswith("✅"):
            c.fill = YES_FILL
            c.font = GREEN_FONT
        elif val.startswith("❌"):
            c.fill = NO_FILL
            c.font = RED_FONT
        elif val.startswith("⚠️"):
            c.fill = PARTIAL_FILL
            c.font = ORANGE_FONT
        else:
            c.fill = ALT_ROW
            c.font = WHITE_FONT
    
    ws.cell(row=row, column=1).fill = ALT_ROW
    ws.cell(row=row, column=2).fill = ALT_ROW
    
    row += 1

# ===== SHEET 2: NeuroCLI Gap Analysis =====
ws2 = wb.create_sheet("Eksik Ozellik Analizi")

ws2.merge_cells('A1:F1')
ws2['A1'] = 'NEUROCLI EKSIK OZELLIK ANALIZI - ONCELIKLENDIRME'
ws2['A1'].font = TITLE_FONT
ws2['A1'].alignment = Alignment(horizontal='center')
ws2.row_dimensions[1].height = 35

gap_headers = ['Oncelik', 'Ozellik', 'Aciklama', 'Rakiplerde Durum', 'Uygulama Zorlugu', 'Etki Puani']
for col, h in enumerate(gap_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = BOLD_WHITE
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = THIN_BORDER

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 14

gaps = [
    ("P0 - Kritik", "MCP Protokol Destegi", "Model Context Protocol ile dis araclar, veritabanlari ve API'lerle baglanti. 3000+ entegrasyon ekosistemi.", "Claude Code, Gemini, OpenCode, Kilo hepsi destekliyor", "Yuksek", "10/10"),
    ("P0 - Kritik", "Gercek Onay Istemi", "Tehlikeli islemler icin kullaniciya gercek soru sorma (su an otomatik onayliyor)", "Tum rakiplerde var", "Dusuk", "9/10"),
    ("P0 - Kritik", "Oturum Devam Ettirme CLI", "/resume veya neuro --continue komutu ile onceki oturumlara donus", "Claude -c, Gemini --resume, OpenCode Ctrl+A", "Dusuk", "9/10"),
    ("P0 - Kritik", "Otomatik Tamamlama", "Slash komutlari, dosya yollari ve model isimleri icin tab-tamamlama", "Claude, Gemini, OpenCode, Kilo hepsinde var", "Orta", "8/10"),
    ("P1 - Yuksek", "VS Code Eklentisi", "VS Code icinde diff goruntuleme, dosya referanslari, context paylasimi", "Claude, Gemini, OpenCode, Kilo hepsinde var", "Yuksek", "8/10"),
    ("P1 - Yuksek", "Diff Onizleme UI", "Dosya degisikliklerini onizleme ve onay mekanizmasi", "Claude inline+IDE, Gemini IDE diff, OpenCode TUI diff", "Orta", "8/10"),
    ("P1 - Yuksek", "Sandbox/Yalitim Sistemi", "Isletim sistemi seviyesinde proses yalitimi (Seatbelt, Bubblewrap)", "Claude Seatbelt+Bubblewrap, Gemini 5 yontem, Kilo Bubblewrap+seccomp", "Yuksek", "7/10"),
    ("P1 - Yuksek", "Ozel Araclar SDK", "Kullanicilarin kendi araclari yazmasi icin SDK/plugin sistemi", "Claude Plugin SDK, OpenCode @opencode-ai/plugin, Kilo custom tools", "Orta", "7/10"),
    ("P1 - Yuksek", "Headless/CI Modu", "CI/CD boru hatlarinda otomatik calisma modu", "Claude --max-turns, Gemini headless, OpenCode -p -f json, Kilo --auto", "Dusuk", "7/10"),
    ("P2 - Orta", "Ozel Ajan Tanimlama (Dosya)", "Markdown/JSON dosyalarla ozel ajan tanimlama", "Claude YAML frontmatter, Gemini agent def files, OpenCode .opencode/agents/", "Dusuk", "6/10"),
    ("P2 - Orta", "Oturum Fork", "Oturumu dallandirma, alternatif yollar deneme", "Claude --fork-session, Gemini /resume save, OpenCode fork, Kilo --fork", "Orta", "6/10"),
    ("P2 - Orta", "Beceri (Skill) Sistemi", "Gorev baglamina gore otomatik aktifolan beciri paketleri", "Claude auto-activate skills, Gemini Agent Skills, OpenCode/Kilo skills/", "Orta", "6/10"),
    ("P2 - Orta", "Harcama Limiti", "Gunluk/harcama limiti ve uyari sistemi", "Claude org limits, Gemini gunluk limit, Kilo per-user daily", "Dusuk", "5/10"),
    ("P2 - Orta", "Prompt Onbellek", "Tekrarlanan promptlar icin onbellek mekanizmasi", "Claude prompt caching, Gemini token caching", "Orta", "5/10"),
    ("P3 - Dusuk", "Tarayici Ajanı", "Web tarayicisi otomasyonu (navigate, form doldurma)", "Claude Chrome ent., Gemini Browser Agent", "Yuksek", "5/10"),
    ("P3 - Dusuk", "Sesli Giris/Cikti", "Ses ile komut verme ve sesli yanit", "Claude voice mode, Kilo voice transcription", "Yuksek", "4/10"),
    ("P3 - Dusuk", "JetBrains Eklentisi", "IntelliJ/PyCharm/WebStorm entegrasyonu", "Claude dedicated, Gemini ACP, Kilo Kotlin plugin", "Yuksek", "4/10"),
    ("P3 - Dusuk", "Yerel Model (Ollama)", "Ollama/LM Studio ile yerel model calistirma", "OpenCode ve Kilo destekliyor", "Orta", "4/10"),
    ("P3 - Dusuk", "Oturum Paylasma", "Oturumlari baskalariyla paylasma", "Claude /export, OpenCode /share, Kilo /share + URL", "Dusuk", "3/10"),
    ("P3 - Dusuk", ".ignore Dosyasi", "AI'in erisemeyecegi dosyalari belirtme", "Gemini .geminiignore, Kilo .kilocodeignore", "Dusuk", "3/10"),
]

CRIT_FILL = PatternFill(start_color="641220", end_color="641220", fill_type="solid")
HIGH_FILL = PatternFill(start_color="7f4f24", end_color="7f4f24", fill_type="solid")
MED_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
LOW_FILL = PatternFill(start_color="1b4332", end_color="1b4332", fill_type="solid")

for i, (pri, feat, desc, rakip, diff, score) in enumerate(gaps, 4):
    ws2.cell(row=i, column=1, value=pri).font = Font(color="FFFFFF", bold=True, size=10)
    ws2.cell(row=i, column=2, value=feat).font = ACCENT_FONT
    ws2.cell(row=i, column=3, value=desc).font = WHITE_FONT
    ws2.cell(row=i, column=4, value=rakip).font = WHITE_FONT
    ws2.cell(row=i, column=5, value=diff).font = WHITE_FONT
    ws2.cell(row=i, column=6, value=score).font = Font(color="E94560", bold=True, size=11)
    
    fill = CRIT_FILL if "P0" in pri else (HIGH_FILL if "P1" in pri else (MED_FILL if "P2" in pri else LOW_FILL))
    for col in range(1, 7):
        ws2.cell(row=i, column=col).fill = fill
        ws2.cell(row=i, column=col).border = THIN_BORDER
        ws2.cell(row=i, column=col).alignment = Alignment(horizontal='center' if col in [1,5,6] else 'left', wrap_text=True)

# ===== SHEET 3: NeuroCLI Unique Strengths =====
ws3 = wb.create_sheet("NeuroCLI Guclu Yonleri")

ws3.merge_cells('A1:D1')
ws3['A1'] = 'NEUROCLI BENZERSIZ AVANTAJLARI - RAKIPLERDE OLMAYAN OZELLIKLER'
ws3['A1'].font = TITLE_FONT
ws3['A1'].alignment = Alignment(horizontal='center')
ws3.row_dimensions[1].height = 35

str_headers = ['Ozellik', 'Detay', 'Rakiplerdeki Durum', 'Avantaj']
for col, h in enumerate(str_headers, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = BOLD_WHITE
    c.fill = NEURO_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = THIN_BORDER

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 45
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 25

strengths = [
    ("23 Ucretsiz Model", "OpenRouter uzerinden 23 tamamen ucretsiz model, gelistirici basina $0 maliyet", "Claude: $4-13/gun, Gemini: sinirli ucretsiz, OpenCode/Kilo: BYOK gerekli", "MALIYET AVANTAJI"),
    ("5 Katmanli Sikistirma", "Tool Budget > Snip > Micro > Session Memory > Full Collapse - en derin sikistirma stratejisi", "Claude: auto-compact sadece, digerleri: tek katmanli", "BAGLAM VERIMLILIGI"),
    ("Danisman Sistemi (Advisor)", "6 farkli tetikleyiciyle ikinci model danismanligi, tekrar eden hata algilama", "Hicbir rakipte yok", "KALITE GARANTISI"),
    ("LSP Entegrasyonu", "TypeScript, Python, Go, Rust icin Language Server Protocol - git-to-definition, references, diagnostics", "Sadece OpenCode'da deneysel, digerlerinde yok", "KOD ZEKASI"),
    ("Repo Haritasi", "5 dilde tanim cikarma, import/reference tespiti, otomatik kod haritasi olusturma", "Sadece Kilo'da codebase indexing var ama daha basit", "PROJE ANLAYISI"),
    ("20 Olay Hook Sistemi", "Session/Agent/Model/Tool/Permission/User/Context/Environment - 8 kategoride 20 lifecycle olayi", "Claude: hooks var ama daha az olay, digerleri: basit hooks", "GENISLETILEBILIRLIK"),
    ("14+ Yerlesik Komut", "init, memory, compact, review, debug, verify, test, refactor, explain, security, perf, migrate, doctor, stats", "Digerleri: 3-8 komut", "VERIMLILIK"),
    ("Ajanlar Arasi Mesajlasma", "5 mesaj tipi (task, result, question, coordination, status) ile derin ajan iletisimi", "Digerleri: sadece sonuc raporu", "KOORDINASYON"),
    ("NEURO.md @import Destegi", "Baska dosyalari @path/to/file.md ile icerme ve 5s polling ile canli guncelleme", "Claude: sadece hiyerarsi, digerleri: statik dosya", "DINAMIK BAGLAM"),
    ("Tehlikeli Komut Algilama", "sudo, rm -rf /, fork bomb, wget|sh gibi 15+ tehlikeli pattern regex ile algilama", "OpenCode: pattern tabanli ama daha az kapsamli", "GUVENLIK"),
]

for i, (feat, detail, rakip, adv) in enumerate(strengths, 4):
    ws3.cell(row=i, column=1, value=feat).font = ACCENT_FONT
    ws3.cell(row=i, column=2, value=detail).font = WHITE_FONT
    ws3.cell(row=i, column=3, value=rakip).font = WHITE_FONT
    ws3.cell(row=i, column=4, value=adv).font = Font(color="52B788", bold=True, size=10)
    for col in range(1, 5):
        ws3.cell(row=i, column=col).fill = ALT_ROW
        ws3.cell(row=i, column=col).border = THIN_BORDER
        ws3.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical='center')

# Save
output_path = "/home/z/my-project/download/NeuroCLI_Kiyaslama_Matrisi.xlsx"
wb.save(output_path)
print(f"XLSX saved to {output_path}")
