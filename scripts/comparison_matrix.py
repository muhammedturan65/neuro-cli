#!/usr/bin/env python3
"""
NeuroCLI vs Competitors - Comprehensive Feature Comparison Matrix 2026
Generates a professional XLSX with multiple sheets covering all features.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

wb = openpyxl.Workbook()
wb.properties.creator = "Z.ai"

# ── Color Palette ──
DRACULA_BG = "#282a36"
DRACULA_PURPLE = "#bd93f9"
DRACULA_GREEN = "#50fa7b"
DRACULA_RED = "#ff5555"
DRACULA_YELLOW = "#f1fa8c"
DRACULA_CYAN = "#8be9fd"
DRACULA_ORANGE = "#ffb86c"
DRACULA_PINK = "#ff79c6"

WHITE = "#FFFFFF"
LIGHT_GRAY = "#f8f9fa"
MED_GRAY = "#e9ecef"
DARK_GRAY = "#495057"
BLACK = "#212529"

# Feature support fills
FILL_YES = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")  # green
FILL_PARTIAL = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")  # yellow
FILL_NO = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")  # red
FILL_NA = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")  # gray

# Header fills
FILL_HEADER = PatternFill(start_color="FFBD93F9", end_color="FFBD93F9", fill_type="solid")
FILL_CATEGORY = PatternFill(start_color="FF282A36", end_color="FF282A36", fill_type="solid")
FILL_NEURO_HEADER = PatternFill(start_color="FF50FA7B", end_color="FF50FA7B", fill_type="solid")
FILL_SCORE = PatternFill(start_color="cfe2ff", end_color="cfe2ff", fill_type="solid")

# Fonts
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
FONT_CATEGORY = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
FONT_NEURO_HEADER = Font(name="Calibri", bold=True, color="FF212529", size=11)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_SCORE = Font(name="Calibri", bold=True, size=12, color="FF0D6EFD")
FONT_TITLE = Font(name="Calibri", bold=True, size=16, color="FFBD93F9")
FONT_SUBTITLE = Font(name="Calibri", bold=True, size=12, color="FF495057")

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Border
THIN_BORDER = Border(
    left=Side(style="thin", color="FFE9ECEF"),
    right=Side(style="thin", color="FFE9ECEF"),
    top=Side(style="thin", color="FFE9ECEF"),
    bottom=Side(style="thin", color="FFE9ECEF")
)

# ── Tools list (10 tools now!) ──
TOOLS = [
    "NeuroCLI",
    "Claude Code",
    "Codex CLI",
    "Aider",
    "Cursor CLI",
    "Gemini CLI",
    "GitHub Copilot CLI",
    "Cline",
    "Goose",
    "OpenHands",
]

# ── Feature Data ──
# Format: (Category, Feature, NeuroCLI, Claude Code, Codex CLI, Aider, Cursor CLI, Gemini CLI, Copilot CLI, Cline, Goose, OpenHands)
# ✅ = Full, ⚠️ = Partial, ❌ = None, N/A = Not Appable

features = [
    # ═══ 1. CORE ARCHITECTURE ═══
    ("Temel Mimari", "CLI / Terminal-Native", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️", "✅", "⚠️"),
    ("Temel Mimari", "Açık Kaynak (OSS)", "✅", "❌", "✅", "✅", "❌", "✅", "❌", "✅", "✅", "✅"),
    ("Temel Mimari", "Çoklu LLM Desteği", "✅", "❌", "❌", "✅", "⚠️", "⚠️", "❌", "✅", "✅", "✅"),
    ("Temel Mimari", "Ücretsiz Model Kullanımı", "✅", "❌", "⚠️", "✅", "❌", "✅", "❌", "✅", "✅", "✅"),
    ("Temel Mimari", "Agentic Davranış (Otonom)", "✅", "✅", "✅", "⚠️", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Temel Mimari", "Multi-Agent Sistem", "✅", "✅", "⚠️", "❌", "✅", "⚠️", "⚠️", "✅", "⚠️", "✅"),
    ("Temel Mimari", "Streaming SSE Çıktı", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Temel Mimari", "Rust / Yüksek Performans", "❌", "❌", "✅", "❌", "❌", "❌", "❌", "❌", "✅", "❌"),
    
    # ═══ 2. CONTEXT & MEMORY ═══
    ("Bağlam & Hafıza", "Bağlam Penceresi Yönetimi", "✅", "✅", "⚠️", "⚠️", "✅", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    ("Bağlam & Hafıza", "Repo Map / Kod Haritası", "✅", "⚠️", "⚠️", "✅", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Bağlam & Hafıza", "NEURO.md / CLAUDE.md / AGENTS.md", "✅", "✅", "✅", "⚠️", "✅", "⚠️", "✅", "✅", "✅", "⚠️"),
    ("Bağlam & Hafıza", ".neuroignore / .gitignore Benzeri", "✅", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Bağlam & Hafıza", "Prompt Cache", "✅", "✅", "⚠️", "⚠️", "✅", "✅", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Bağlam & Hafıza", "Tree-sitter Entegrasyonu", "❌", "❌", "❌", "✅", "❌", "❌", "❌", "❌", "❌", "❌"),
    ("Bağlam & Hafıza", "Otomatik Bağlam Sıkıştırma", "✅", "✅", "⚠️", "⚠️", "✅", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    
    # ═══ 3. TOOL SYSTEM ═══
    ("Araç Sistemi", "Dosya Okuma/Yazma", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Araç Sistemi", "Bash / Shell Komut Çalıştırma", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Araç Sistemi", "Web Arama / Tarama", "✅", "⚠️", "⚠️", "❌", "⚠️", "✅", "⚠️", "✅", "⚠️", "✅"),
    ("Araç Sistemi", "MCP Protokol Desteği", "✅", "✅", "✅", "❌", "✅", "✅", "✅", "✅", "✅", "⚠️"),
    ("Araç Sistemi", "Özel Araç Tanımlama (Custom Tools)", "✅", "⚠️", "⚠️", "❌", "⚠️", "✅", "⚠️", "✅", "✅", "⚠️"),
    ("Araç Sistemi", "Plugin SDK Sistemi", "✅", "✅", "✅", "⚠️", "⚠️", "✅", "✅", "⚠️", "✅", "⚠️"),
    ("Araç Sistemi", "Function Calling / Tool Use", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Araç Sistemi", "Tarayıcı Otomasyonu", "❌", "⚠️", "❌", "❌", "⚠️", "❌", "⚠️", "✅", "❌", "✅"),
    
    # ═══ 4. SESSION & WORKFLOW ═══
    ("Oturum & İş Akışı", "Oturum Yönetimi (Resume/Save)", "✅", "✅", "⚠️", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Oturum & İş Akışı", "Oturum Etiketleme (Tags)", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Oturum & İş Akışı", "Undo / Redo Sistemi", "✅", "✅", "⚠️", "✅", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Oturum & İş Akışı", "Git Checkpoint / Otomatik Commit", "✅", "✅", "✅", "✅", "⚠️", "⚠️", "✅", "✅", "⚠️", "⚠️"),
    ("Oturum & İş Akışı", "Diff Önizleme (Değişiklik Görüntüleme)", "✅", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅", "⚠️", "⚠️"),
    ("Oturum & İş Akışı", "Plan / Act Modu", "⚠️", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅", "⚠️", "⚠️"),
    ("Oturum & İş Akışı", "Headless Mod (Betik Modu)", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅"),
    ("Oturum & İş Akışı", "Planlı Görevler / Scheduled Tasks", "❌", "✅", "❌", "❌", "✅", "❌", "⚠️", "❌", "❌", "❌"),
    
    # ═══ 5. SAFETY & APPROVAL ═══
    ("Güvenlik & Onay", "İnteraktif Onay Sistemi", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Güvenlik & Onay", "Sandbox İzolasyonu", "✅", "⚠️", "✅", "❌", "⚠️", "⚠️", "✅", "⚠️", "⚠️", "✅"),
    ("Güvenlik & Onay", "Kernel-Level Sandbox", "❌", "❌", "✅", "❌", "❌", "❌", "❌", "❌", "❌", "⚠️"),
    ("Güvenlik & Onay", "Doom Loop Önleme", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Güvenlik & Onay", "Harcanan Para İzleme", "✅", "⚠️", "⚠️", "⚠️", "✅", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    ("Güvenlik & Onay", "Güvenlik Danışmanı (Advisor)", "✅", "⚠️", "⚠️", "❌", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    
    # ═══ 6. SKILL SYSTEM ═══
    ("Yetenek (Skill) Sistemi", "SKILL.md Standardı (YAML Frontmatter)", "⚠️", "✅", "✅", "❌", "⚠️", "✅", "✅", "⚠️", "⚠️", "❌"),
    ("Yetenek (Skill) Sistemi", "Otomatik Skill Keşfi / Aktivasyon", "⚠️", "✅", "✅", "❌", "⚠️", "✅", "✅", "⚠️", "⚠️", "❌"),
    ("Yetenek (Skill) Sistemi", "Skill Paylaşımı (GitHub Registry)", "❌", "✅", "✅", "❌", "⚠️", "✅", "✅", "❌", "❌", "❌"),
    ("Yetenek (Skill) Sistemi", "Özel Ajanlar (Custom Agents)", "✅", "✅", "⚠️", "❌", "⚠️", "✅", "✅", "✅", "⚠️", "⚠️"),
    ("Yetenek (Skill) Sistemi", "AGENTS.md Evrensel Standardı", "⚠️", "✅", "✅", "❌", "✅", "⚠️", "✅", "✅", "✅", "❌"),
    ("Yetenek (Skill) Sistemi", "agentskills.io Spec Uyumu", "❌", "✅", "✅", "❌", "⚠️", "⚠️", "✅", "❌", "❌", "❌"),
    
    # ═══ 7. HOOKS & LIFECYCLE ═══
    ("Hook & Yaşam Döngüsü", "Lifecycle Hooks Sistemi", "✅", "✅", "⚠️", "❌", "⚠️", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    ("Hook & Yaşam Döngüsü", "20+ Hook Olayı", "✅", "✅", "⚠️", "❌", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Hook & Yaşam Döngüsü", "Plugin Bundle (Skills+Hooks+MCP)", "⚠️", "✅", "⚠️", "❌", "⚠️", "✅", "✅", "⚠️", "⚠️", "⚠️"),
    
    # ═══ 8. MODEL MANAGEMENT ═══
    ("Model Yönetimi", "Model Router (Akıllı Yönlendirme)", "✅", "⚠️", "⚠️", "⚠️", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Model Yönetimi", "Fallback Zinciri", "✅", "⚠️", "⚠️", "⚠️", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Model Yönetimi", "Ollama / Yerel Model Desteği", "✅", "❌", "❌", "✅", "❌", "❌", "❌", "✅", "✅", "✅"),
    ("Model Yönetimi", "Extended Thinking (Derin Düşünme)", "✅", "✅", "⚠️", "❌", "⚠️", "✅", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Model Yönetimi", "Token Kullanım Takibi", "✅", "✅", "✅", "⚠️", "✅", "✅", "✅", "⚠️", "⚠️", "⚠️"),
    ("Model Yönetimi", "Maliyet Hesaplama", "✅", "✅", "⚠️", "⚠️", "✅", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    
    # ═══ 9. UX & THEME ═══
    ("UX & Tema", "Tema Sistemi (4+ Tema)", "✅", "⚠️", "❌", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("UX & Tema", "Tab Tamamlama (35+ Komut)", "✅", "✅", "⚠️", "✅", "✅", "✅", "✅", "⚠️", "⚠️", "⚠️"),
    ("UX & Tema", "Shell Completion (Bash/Zsh/Fish)", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️", "⚠️", "⚠️"),
    ("UX & Tema", "Vim Keybinding Modu", "✅", "❌", "❌", "⚠️", "❌", "❌", "❌", "❌", "❌", "❌"),
    ("UX & Tema", "Çıktı Stili Seçenekleri", "✅", "⚠️", "⚠️", "⚠️", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("UX & Tema", "Markdown Render", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    
    # ═══ 10. MULTIMODAL & VOICE ═══
    ("Çoklu Ortam & Ses", "Görsel / Resim Desteği", "✅", "✅", "✅", "⚠️", "✅", "✅", "✅", "✅", "⚠️", "⚠️"),
    ("Çoklu Ortam & Ses", "Ses Girişi (STT)", "✅", "⚠️", "❌", "✅", "❌", "❌", "❌", "❌", "❌", "❌"),
    ("Çoklu Ortam & Ses", "Ses Çıkışı (TTS)", "✅", "⚠️", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "❌"),
    
    # ═══ 11. I18N & ACCESSIBILITY ═══
    ("Uluslararasılaşma", "Çoklu Dil Desteği (i18n)", "✅", "❌", "❌", "❌", "❌", "⚠️", "❌", "❌", "❌", "❌"),
    ("Uluslararasılaşma", "Türkçe Dil Desteği", "✅", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "❌"),
    ("Uluslararasılaşma", "Sağdan Sola (RTL) Desteği", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "❌"),
    
    # ═══ 12. API & CLOUD ═══
    ("API & Bulut", "API Server Modu (REST+WS)", "✅", "❌", "❌", "❌", "⚠️", "❌", "⚠️", "❌", "❌", "⚠️"),
    ("API & Bulut", "Bulut Senkronizasyon (Gist)", "✅", "⚠️", "❌", "❌", "✅", "❌", "✅", "❌", "❌", "❌"),
    ("API & Bulut", "Web Dashboard", "✅", "❌", "❌", "❌", "✅", "❌", "✅", "❌", "❌", "✅"),
    ("API & Bulut", "Cloud Agent (Bulutta Çalışma)", "❌", "❌", "✅", "❌", "✅", "❌", "✅", "❌", "❌", "✅"),
    ("API & Bulut", "Plan-to-Cloud Handoff", "❌", "❌", "❌", "❌", "✅", "❌", "❌", "❌", "❌", "❌"),
    
    # ═══ 13. INTEGRATIONS ═══
    ("Entegrasyonlar", "LSP Entegrasyonu", "✅", "⚠️", "❌", "❌", "✅", "❌", "⚠️", "❌", "❌", "❌"),
    ("Entegrasyonlar", "Git Entegrasyonu (Derin)", "✅", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅", "⚠️", "⚠️"),
    ("Entegrasyonlar", "GitHub PR/Issue İşlemleri", "⚠️", "⚠️", "⚠️", "❌", "✅", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    ("Entegrasyonlar", "IDE Entegrasyonu (VS Code vb.)", "❌", "⚠️", "✅", "⚠️", "✅", "❌", "✅", "✅", "✅", "✅"),
    ("Entegrasyonlar", "CI/CD Pipeline Entegrasyonu", "❌", "✅", "✅", "⚠️", "✅", "✅", "✅", "⚠️", "⚠️", "⚠️"),
    ("Entegrasyonlar", "Jira/Linear Entegrasyonu", "❌", "⚠️", "⚠️", "❌", "⚠️", "⚠️", "✅", "❌", "❌", "❌"),
    
    # ═══ 14. MONITORING & TELEMETRY ═══
    ("İzleme & Telemetri", "Anonim Telemetri (Opt-in)", "✅", "⚠️", "⚠️", "❌", "⚠️", "⚠️", "⚠️", "❌", "❌", "❌"),
    ("İzleme & Telemetri", "Kullanım İstatistikleri", "✅", "✅", "⚠️", "⚠️", "✅", "✅", "✅", "⚠️", "⚠️", "✅"),
    ("İzleme & Telemetri", "Performans Metrikleri", "✅", "⚠️", "⚠️", "❌", "✅", "⚠️", "✅", "⚠️", "⚠️", "✅"),
    
    # ═══ 15. CODE QUALITY ═══
    ("Kod Kalitesi", "Linting Entegrasyonu", "❌", "⚠️", "⚠️", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
    ("Kod Kalitesi", "Test Çalıştırma Entegrasyonu", "❌", "✅", "✅", "✅", "✅", "⚠️", "✅", "⚠️", "⚠️", "✅"),
    ("Kod Kalitesi", "Kod İnceleme (Review)", "❌", "✅", "✅", "❌", "✅", "⚠️", "✅", "⚠️", "⚠️", "⚠️"),
    ("Kod Kalitesi", "Refactoring Asistanı", "⚠️", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅", "⚠️", "⚠️"),
    ("Kod Kalitesi", "Güvenlik Tarama", "❌", "⚠️", "⚠️", "❌", "⚠️", "⚠️", "✅", "❌", "❌", "⚠️"),
    
    # ═══ 16. DEVELOPER EXPERIENCE ═══
    ("Geliştirici Deneyimi", "Kurulum Kolaylığı (1 Komut)", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️"),
    ("Geliştirici Deneyimi", "Dokümantasyon Kalitesi", "⚠️", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅"),
    ("Geliştirici Deneyimi", "Topluluk Büyüklüğü", "⚠️", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️", "✅"),
    ("Geliştirici Deneyimi", "VS Code Uzantısı", "❌", "❌", "✅", "⚠️", "✅", "❌", "✅", "✅", "✅", "⚠️"),
    ("Geliştirici Deneyimi", "Masaüstü Uygulaması", "❌", "❌", "❌", "❌", "✅", "❌", "❌", "❌", "✅", "✅"),
    
    # ═══ 17. ADVANCED AI FEATURES ═══
    ("Gelişmiş AI Özellikleri", "Auto Mod (Tam Otonom)", "❌", "✅", "✅", "❌", "✅", "❌", "✅", "⚠️", "⚠️", "⚠️"),
    ("Gelişmiş AI Özellikleri", "/goal / /routine Komutları", "❌", "✅", "❌", "❌", "⚠️", "❌", "⚠️", "❌", "❌", "❌"),
    ("Gelişmiş AI Özellikleri", "Paralel Agent Çalıştırma", "❌", "✅", "⚠️", "❌", "✅", "⚠️", "⚠️", "✅", "⚠️", "✅"),
    ("Gelişmiş AI Özellikleri", "Background Session", "❌", "✅", "❌", "❌", "✅", "❌", "⚠️", "❌", "❌", "⚠️"),
    ("Gelişmiş AI Özellikleri", "Retry / Self-Healing Mantığı", "✅", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅", "⚠️", "✅"),
    ("Gelişmiş AI Özellikleri", "SWE-bench Skoru", "❌", "✅", "✅", "✅", "✅", "⚠️", "⚠️", "⚠️", "⚠️", "⚠️"),
]

# ── Helper Functions ──
def score_cell(val):
    """Convert symbol to numeric score"""
    if val == "✅": return 3
    if val == "⚠️": return 1.5
    if val == "❌": return 0
    return 0

def fill_for_value(val):
    if val == "✅": return FILL_YES
    if val == "⚠️": return FILL_PARTIAL
    if val == "❌": return FILL_NO
    return FILL_NA

def create_matrix_sheet(ws, title):
    """Create the main feature comparison matrix sheet"""
    ws.title = title
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1, value=title).font = FONT_TITLE
    ws.cell(row=1, column=1).alignment = ALIGN_LEFT
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(row=2, column=1, value="NeuroCLI Kapsamlı Rakip Analizi - Temmuz 2026 | 10 Araç × 75+ Özellik").font = FONT_SUBTITLE
    ws.cell(row=2, column=1).alignment = ALIGN_LEFT
    
    # Headers
    row = 4
    headers = ["Kategori", "Özellik"] + TOOLS
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER if col > 2 else FONT_CATEGORY
        cell.fill = FILL_HEADER if col > 2 else FILL_CATEGORY
        if h == "NeuroCLI":
            cell.fill = FILL_NEURO_HEADER
            cell.font = FONT_NEURO_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Data rows
    current_category = ""
    for i, feat in enumerate(features):
        row = 5 + i
        cat, name = feat[0], feat[1]
        values = feat[2:]
        
        # Category column
        cell_cat = ws.cell(row=row, column=1, value=cat)
        cell_feat = ws.cell(row=row, column=2, value=name)
        
        if cat != current_category:
            current_category = cat
            cell_cat.font = FONT_BOLD
        else:
            cell_cat.font = FONT_NORMAL
        
        cell_feat.font = FONT_NORMAL
        cell_cat.alignment = ALIGN_LEFT
        cell_feat.alignment = ALIGN_LEFT
        cell_cat.border = THIN_BORDER
        cell_feat.border = THIN_BORDER
        
        # Tool columns
        for col_idx, val in enumerate(values, 3):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.fill = fill_for_value(val)
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
    
    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 38
    for col_idx in range(3, 13):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    
    # Row height
    for r in range(4, 5 + len(features)):
        ws.row_dimensions[r].height = 22
    
    return ws


def create_score_sheet(ws):
    """Create the scoring summary sheet"""
    ws.title = "Skor Tablosu"
    
    # Calculate scores by category
    categories = {}
    for feat in features:
        cat = feat[0]
        if cat not in categories:
            categories[cat] = {t: 0 for t in TOOLS}
        for i, t in enumerate(TOOLS):
            categories[cat][t] += score_cell(feat[2 + i])
    
    # Total scores
    totals = {t: 0 for t in TOOLS}
    for cat_data in categories.values():
        for t in TOOLS:
            totals[t] += cat_data[t]
    
    # Max possible score
    max_per_feature = 3
    total_features = len(features)
    max_score = max_per_feature * total_features
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1, value="NeuroCLI Skor Karşılaştırma Tablosu").font = FONT_TITLE
    ws.cell(row=1, column=1).alignment = ALIGN_LEFT
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(row=2, column=1, value=f"Maksimum Mümkün Skor: {max_score} | Toplam Özellik: {total_features}").font = FONT_SUBTITLE
    
    # Headers
    row = 4
    headers = ["Kategori"] + TOOLS
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER if col > 1 else FONT_CATEGORY
        cell.fill = FILL_HEADER if col > 1 else FILL_CATEGORY
        if h == "NeuroCLI":
            cell.fill = FILL_NEURO_HEADER
            cell.font = FONT_NEURO_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Category scores
    for i, (cat, scores) in enumerate(categories.items()):
        row = 5 + i
        ws.cell(row=row, column=1, value=cat).font = FONT_BOLD
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).border = THIN_BORDER
        for col_idx, t in enumerate(TOOLS, 2):
            cell = ws.cell(row=row, column=col_idx, value=scores[t])
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            # Color scale
            pct = scores[t] / (max_per_feature * sum(1 for f in features if f[0] == cat))
            if pct >= 0.7:
                cell.fill = FILL_YES
            elif pct >= 0.4:
                cell.fill = FILL_PARTIAL
            else:
                cell.fill = FILL_NO
    
    # Total row
    total_row = 5 + len(categories)
    ws.cell(row=total_row, column=1, value="TOPLAM SKOR").font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=total_row, column=1).fill = FILL_SCORE
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for col_idx, t in enumerate(TOOLS, 2):
        cell = ws.cell(row=total_row, column=col_idx, value=totals[t])
        cell.font = FONT_SCORE
        cell.fill = FILL_SCORE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Percentage row
    pct_row = total_row + 1
    ws.cell(row=pct_row, column=1, value="YÜZDE (%)").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=pct_row, column=1).fill = FILL_SCORE
    ws.cell(row=pct_row, column=1).border = THIN_BORDER
    for col_idx, t in enumerate(TOOLS, 2):
        pct = round((totals[t] / max_score) * 100, 1)
        cell = ws.cell(row=pct_row, column=col_idx, value=f"%{pct}")
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.fill = FILL_SCORE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Rank row
    rank_row = pct_row + 1
    ws.cell(row=rank_row, column=1, value="SIRALAMA").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=rank_row, column=1).fill = FILL_SCORE
    ws.cell(row=rank_row, column=1).border = THIN_BORDER
    sorted_tools = sorted(TOOLS, key=lambda t: totals[t], reverse=True)
    for col_idx, t in enumerate(TOOLS, 2):
        rank = sorted_tools.index(t) + 1
        cell = ws.cell(row=rank_row, column=col_idx, value=f"#{rank}")
        cell.font = Font(name="Calibri", bold=True, size=12, color="FF0D6EFD" if rank == 1 else "198754" if rank <= 3 else "6c757d")
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if rank == 1:
            cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 28
    for col_idx in range(2, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    
    return totals, max_score


def create_gaps_sheet(ws):
    """Create the gap analysis sheet - what NeuroCLI is missing"""
    ws.title = "Eksik Tespiti"
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value="NeuroCLI Eksik Özellik Tespiti").font = FONT_TITLE
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=2, column=1, value="Rakiplerde var ama NeuroCLI'de eksik olan özellikler ve öncelik seviyeleri").font = FONT_SUBTITLE
    
    # Headers
    headers = ["#", "Kategori", "Eksik Özellik", "En İyi Rakip", "Öncelik", "Etki Skoru", "Uygulama Zorluğu", "Açıklama"]
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Gaps data
    gaps = [
        (1, "Yetenek Sistemi", "SKILL.md YAML Frontmatter Standardı", "Claude Code, Codex CLI", "P0 - Kritik", 10, "Orta", "agentskills.io spesifikasyonu: YAML frontmatter ile name, description, triggers alanları. Skill keşfi ve aktivasyonu için zorunlu."),
        (2, "Yetenek Sistemi", "Otomatik Skill Keşfi ve Aktivasyonu", "Claude Code, Codex CLI", "P0 - Kritik", 10, "Orta", "Ajan başlangıcında skill'lerin name/description alanlarını okuyup otomatik eşleştirme. Mevcut skill-system.ts bu mekanizmadan yoksun."),
        (3, "Yetenek Sistemi", "Skill Paylaşım Registry (GitHub)", "Claude Code, Copilot CLI", "P0 - Kritik", 9, "Yüksek", "anthropics/skills benzeri merkezi skill deposu. Kullanıcılar skill yükleyip paylaşabilmeli."),
        (4, "Yetenek Sistemi", "agentskills.io Spec Uyumu", "Claude Code, Codex CLI, Copilot CLI", "P0 - Kritik", 9, "Orta", "Evrensel SKILL.md spesifikasyonu ile uyumluluk. Frontmatter formatı, dizin yapısı ve keşif mekanizması."),
        (5, "Gelişmiş AI", "Auto Mod (Tam Otonom Çalışma)", "Claude Code, Codex CLI, Cursor CLI", "P0 - Kritik", 10, "Yüksek", "Onay istemeden tam otonom çalışma modu. Claude Code Auto Mode + /goal + /routine benzeri."),
        (6, "Gelişmiş AI", "Planlı Görevler / Scheduled Tasks (/loop)", "Claude Code", "P0 - Kritik", 9, "Yüksek", "Periyodik görev zamanlama. Claude Code /loop komutu ile tekrarlayan görevleri otomatik çalıştırma."),
        (7, "Gelişmiş AI", "Paralel Agent Çalıştırma", "Claude Code, Cursor CLI", "P1 - Önemli", 8, "Yüksek", "Birden fazla agent'ı aynı anda farklı görevlerde çalıştırabilme. Sub-agent paralel execution."),
        (8, "Gelişmiş AI", "Background Session", "Claude Code, Cursor CLI", "P1 - Önemli", 8, "Yüksek", "Arka planda çalışan oturumlar. Agent View ile oturumları izleyebilme."),
        (9, "Temel Mimari", "Rust / Yüksek Performans Çekirdek", "Codex CLI, Goose", "P2 - Orta", 6, "Çok Yüksek", "TypeScript yerine Rust ile kritik yolların yeniden yazımı. Büyük repolarda performans avantajı."),
        (10, "Araç Sistemi", "Tarayıcı Otomasyonu (Puppeteer/Playwright)", "Cline, OpenHands", "P1 - Önemli", 8, "Yüksek", "Web sayfalarında otomatik tıklama, form doldurma, ekran görüntüsü alma. Cline browser tool."),
        (11, "Entegrasyonlar", "IDE Entegrasyonu (VS Code Extension)", "Cursor, Cline, Goose, Codex", "P1 - Önemli", 9, "Yüksek", "VS Code uzantısı olarak çalışabilme. En çok kullanılan geliştirme ortamı."),
        (12, "Entegrasyonlar", "CI/CD Pipeline Entegrasyonu", "Claude Code, Cursor, Copilot CLI", "P1 - Önemli", 7, "Orta", "GitHub Actions, GitLab CI gibi pipeline'larda agent çalıştırabilme."),
        (13, "Entegrasyonlar", "GitHub PR/Issue İşlemleri", "Cursor CLI, Copilot CLI", "P1 - Önemli", 7, "Orta", "PR açma, inceleme, issue yönetimi gibi GitHub iş akışı entegrasyonu."),
        (14, "Kod Kalitesi", "Linting Entegrasyonu (Otomatik)", "Aider", "P1 - Önemli", 7, "Düşük", "ESLint, Pylint vb. entegrasyonu. Kod değişikliği sonrası otomatik lint çalıştırma."),
        (15, "Kod Kalitesi", "Test Çalıştırma Entegrasyonu", "Claude Code, Codex CLI, Aider", "P1 - Önemli", 8, "Orta", "Değişiklik sonrası test suite'i otomatik çalıştırma ve hata düzeltme döngüsü."),
        (16, "Kod Kalitesi", "Kod İnceleme (Code Review)", "Claude Code, Codex CLI, Cursor", "P1 - Önemli", 7, "Orta", "PR diff'lerini otomatik inceleyip geri bildirim verme."),
        (17, "Kod Kalitesi", "Güvenlik Tarama", "Copilot CLI", "P2 - Orta", 6, "Orta", "Otomatik güvenlik açığı tarama ve öneriler."),
        (18, "Güvenlik", "Kernel-Level Sandbox", "Codex CLI", "P2 - Orta", 7, "Çok Yüksek", "Rust tabanlı kernel izolasyonu. Codex CLI'nin en güçlü güvenlik özelliği."),
        (19, "Hook & Yaşam Döngüsü", "Plugin Bundle (Skills+Hooks+MCP Tek Paket)", "Claude Code, Copilot CLI", "P1 - Önemli", 8, "Orta", "Skill, hook ve MCP sunucusunu tek kurulabilir paket olarak sunma."),
        (20, "Bağlam & Hafıza", "Tree-sitter Entegrasyonu", "Aider", "P1 - Önemli", 8, "Yüksek", "Kod tabanını AST seviyesinde analiz. Repo map oluşturma. Aider'ın en güçlü özelliği."),
        (21, "API & Bulut", "Cloud Agent (Bulutta Çalışma)", "Codex CLI, Cursor CLI, Copilot CLI", "P1 - Önemli", 8, "Çok Yüksek", "Bulut sanal ortamında agent çalıştırma. Kullanıcı makinesi dışında izole çalışma."),
        (22, "API & Bulut", "Plan-to-Cloud Handoff", "Cursor CLI", "P2 - Orta", 6, "Yüksek", "Yerel planlamayı bulut ortamına devredip arka planda çalıştırma."),
        (23, "Geliştirici Deneyimi", "VS Code Uzantısı", "Cline, Cursor, Codex", "P1 - Önemli", 9, "Yüksek", "IDE içinde terminal aracı olarak çalışabilme."),
        (24, "Gelişmiş AI", "SWE-bench Benchmark Skoru", "Claude Code, Codex CLI, Aider", "P2 - Orta", 5, "Çok Yüksek", "Endüstri standardı benchmark'ta rekabetçi skor elde etme."),
        (25, "Entegrasyonlar", "Jira/Linear Proje Yönetimi", "Copilot CLI", "P2 - Orta", 5, "Orta", "Proje yönetim araçlarıyla entegrasyon."),
        (26, "Uluslararasılaşma", "RTL (Sağdan Sola) Dil Desteği", "Yok", "P3 - Düşük", 3, "Orta", "Arapça, İbranice gibi RTL diller için destek."),
    ]
    
    for i, gap in enumerate(gaps):
        row = 5 + i
        for col, val in enumerate(gap, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if col > 3 else ALIGN_CENTER
            cell.border = THIN_BORDER
            
            # Priority coloring
            if col == 5:
                if "P0" in str(val):
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF842029")
                elif "P1" in str(val):
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF664D03")
                elif "P2" in str(val):
                    cell.fill = PatternFill(start_color="cfe2ff", end_color="cfe2ff", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF084298")
                else:
                    cell.fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF495057")
    
    # Column widths
    widths = [5, 22, 42, 28, 16, 12, 18, 70]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    
    return ws


def create_skills_md_sheet(ws):
    """Create the SKILL.md ecosystem analysis sheet"""
    ws.title = "Skill Ecosystem"
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value="SKILL.md Ekosistem Analizi").font = FONT_TITLE
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=2, column=1, value="Bağımsız skill/md dosyaları ve GitHub'daki paylaşım ekosistemi").font = FONT_SUBTITLE
    
    # Headers
    headers = ["#", "Araç / Standart", "Dosya Formatı", "YAML Frontmatter", "Otomatik Keşif", "GitHub Registry", "Cross-Tool Uyum", "Detay"]
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Skills ecosystem data
    skills_data = [
        (1, "Claude Code Skills", ".claude/skills/<name>/SKILL.md", "✅ name + description", "✅ Ajan başlangıcında", "✅ anthropics/skills repo", "✅ Codex, Copilot", "Anthropic'ın resmi skill standardı. 345+ topluluk skill'i mevcut."),
        (2, "Codex CLI AGENTS.md", "AGENTS.md (repo root)", "❌ Sadece markdown", "✅ Otomatik okuma", "❌ Resmi registry yok", "✅ Çoğu araç okur", "OpenAI'nin proje talimat formatı. Evrensel proje seviyesi talimatlar."),
        (3, "SKILL.md (agentskills.io)", "<skill-name>/SKILL.md", "✅ name + description (zorunlu)", "✅ Frontmatter tarama", "✅ agentskills.io", "✅ Açık standart", "Açık spesifikasyon: YAML frontmatter + markdown body. Her araç uyumlu."),
        (4, ".cursorrules", ".cursor/rules (repo root)", "❌ Sadece markdown", "✅ Otomatik okuma", "⚠️ Topluluk listeleri", "❌ Cursor-özel", "Cursor IDE'ye özel talimat formatı. CLI'da da destekleniyor."),
        (5, "CLAUDE.md", "CLAUDE.md (repo root)", "❌ Sadece markdown", "✅ Otomatik okuma", "❌ Resmi registry yok", "⚠️ Claude Code öncelikli", "Claude Code'un proje talimat dosyası. AGENTS.md ile birleşiyor."),
        (6, ".clinerules", ".clinerules (repo root)", "❌ Sadece markdown", "✅ Otomatik okuma", "❌ Resmi registry yok", "❌ Cline-özel", "Cline'ın proje seviyesi kural dosyası."),
        (7, ".goosehints", ".goosehints (repo root)", "❌ Sadece markdown", "✅ Otomatik okuma", "❌ Resmi registry yok", "❌ Goose-özel", "Goose agent için ipucu ve talimat dosyası."),
        (8, "Copilot Skills", ".github/skills/<name>/SKILL.md", "✅ name + description", "✅ Ajan başlangıcında", "✅ GitHub Marketplace", "✅ SKILL.md uyumlu", "GitHub Copilot'un skill sistemi. SKILL.md standardını benimsedi."),
        (9, "NeuroCLI Skills", ".neuro/skills/<name>/skill.md", "⚠️ Basit format", "⚠️ Kısmi", "❌ Henüz yok", "❌ NeuroCLI-özel", "Mevcut implementasyon. SKILL.md standardına uyumlu hale getirilmeli."),
    ]
    
    for i, data in enumerate(skills_data):
        row = 5 + i
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER if col <= 7 else ALIGN_LEFT
            cell.border = THIN_BORDER
            if isinstance(val, str) and val in ("✅", "❌", "⚠️"):
                cell.fill = fill_for_value(val)
    
    # Column widths
    widths = [5, 24, 30, 22, 18, 22, 22, 55]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    
    return ws


def create_chart_sheet(ws, totals, max_score):
    """Create a chart sheet with bar chart comparison"""
    ws.title = "Görsel Karşılaştırma"
    
    # Title
    ws.cell(row=1, column=1, value="NeuroCLI vs Rakipler - Görsel Skor Karşılaştırması").font = FONT_TITLE
    
    # Data table for chart
    ws.cell(row=3, column=1, value="Araç").font = FONT_BOLD
    ws.cell(row=3, column=2, value="Toplam Skor").font = FONT_BOLD
    ws.cell(row=3, column=3, value="Yüzde (%)").font = FONT_BOLD
    
    sorted_tools = sorted(TOOLS, key=lambda t: totals[t], reverse=True)
    for i, t in enumerate(sorted_tools):
        row = 4 + i
        ws.cell(row=row, column=1, value=t).font = FONT_NORMAL
        ws.cell(row=row, column=2, value=totals[t]).font = FONT_NORMAL
        pct = round((totals[t] / max_score) * 100, 1)
        ws.cell(row=row, column=3, value=pct).font = FONT_NORMAL
    
    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "AI Terminal Araçları - Toplam Özellik Skoru"
    chart.y_axis.title = "Skor"
    chart.x_axis.title = "Araçlar"
    chart.width = 28
    chart.height = 16
    
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(TOOLS), max_col=2)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(TOOLS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    
    # Data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    
    ws.add_chart(chart, "A16")
    
    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    return ws


def create_priority_sheet(ws):
    """Create the priority-based implementation roadmap"""
    ws.title = "Uygulama Yol Haritası"
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(row=1, column=1, value="NeuroCLI Uygulama Yol Haritası - Öncelik Sıralaması").font = FONT_TITLE
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws.cell(row=2, column=1, value="P0 = Hemen | P1 = Kısa Vadeli | P2 = Orta Vadeli | P3 = Uzun Vadeli").font = FONT_SUBTITLE
    
    # Headers
    headers = ["#", "Öncelik", "Özellik", "Kategori", "Etki (1-10)", "Zorluk", "Tahmini Süre", "Bağımlılıklar", "Açıklama"]
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    
    # Roadmap data
    roadmap = [
        (1, "P0", "SKILL.md Standardı Uyumu", "Yetenek Sistemi", 10, "Orta", "2-3 gün", "Yok", "YAML frontmatter formatı, dizin yapısı, agentskills.io spec ile tam uyumluluk"),
        (2, "P0", "Otomatik Skill Keşfi", "Yetenek Sistemi", 10, "Orta", "2-3 gün", "SKILL.md Standardı", "Ajan başlangıcında skill frontmatter'larını okuyup otomatik eşleştirme"),
        (3, "P0", "Skill Registry / Marketplace", "Yetenek Sistemi", 9, "Yüksek", "5-7 gün", "SKILL.md Standardı", "Merkezi skill deposu. neuro skill install/search/publish komutları"),
        (4, "P0", "Auto Mod (Tam Otonom)", "Gelişmiş AI", 10, "Yüksek", "3-5 gün", "Yok", "Onay istemeden tam otonom çalışma. /auto komutu + güvenlik katmanı"),
        (5, "P0", "Scheduled Tasks (/loop)", "Gelişmiş AI", 9, "Yüksek", "3-5 gün", "Auto Mod", "Periyodik görev zamanlama. /loop <interval> <prompt> komutu"),
        (6, "P1", "Paralel Agent Çalıştırma", "Gelişmiş AI", 8, "Yüksek", "5-7 gün", "Multi-Agent Sistemi", "Birden fazla agent'ı aynı anda çalıştırma. Sub-agent paralel execution"),
        (7, "P1", "Background Session", "Gelişmiş AI", 8, "Yüksek", "3-5 gün", "Session Yönetimi", "Arka planda çalışan oturumlar. /bg <prompt> ile başlatma"),
        (8, "P1", "Tarayıcı Otomasyonu", "Araç Sistemi", 8, "Yüksek", "5-7 gün", "Yok", "Puppeteer/Playwright ile web sayfası otomasyonu. Screen capture, click, type"),
        (9, "P1", "Tree-sitter Entegrasyonu", "Bağlam & Hafıza", 8, "Yüksek", "5-7 gün", "Yok", "AST seviyesinde kod analizi. Repo map oluşturma. Aider tarzı"),
        (10, "P1", "VS Code Uzantısı", "Entegrasyonlar", 9, "Yüksek", "10-15 gün", "API Server Modu", "VS Code içinde terminal aracı olarak çalışma. Webview panel"),
        (11, "P1", "Linting + Test Entegrasyonu", "Kod Kalitesi", 8, "Orta", "3-5 gün", "Yok", "Değişiklik sonrası otomatik lint ve test çalıştırma döngüsü"),
        (12, "P1", "Code Review Modu", "Kod Kalitesi", 7, "Orta", "3-5 gün", "Linting + Test", "PR diff'lerini otomatik inceleyip geri bildirim verme"),
        (13, "P1", "GitHub PR/Issue İşlemleri", "Entegrasyonlar", 7, "Orta", "3-5 gün", "Git Entegrasyonu", "PR açma, inceleme, issue yönetimi"),
        (14, "P1", "Plugin Bundle Sistemi", "Hook & Yaşam Döngüsü", 8, "Orta", "3-5 gün", "Skills + Hooks + MCP", "Skill, hook ve MCP'yi tek pakette birleştirme"),
        (15, "P1", "CI/CD Pipeline Entegrasyonu", "Entegrasyonlar", 7, "Orta", "3-5 gün", "Headless Mod", "GitHub Actions, GitLab CI'da agent çalıştırma"),
        (16, "P2", "Kernel-Level Sandbox", "Güvenlik", 7, "Çok Yüksek", "15-20 gün", "Rust bilgisi", "Codex CLI tarzı kernel izolasyonu"),
        (17, "P2", "Cloud Agent", "API & Bulut", 8, "Çok Yüksek", "15-20 gün", "Bulut altyapısı", "Bulut sanal ortamında agent çalıştırma"),
        (18, "P2", "Rust Performans Çekirdek", "Temel Mimari", 6, "Çok Yüksek", "30+ gün", "Rust bilgisi", "Kritik yolların Rust ile yeniden yazımı"),
        (19, "P2", "Güvenlik Tarama", "Kod Kalitesi", 6, "Orta", "5-7 gün", "Yok", "Otomatik güvenlik açığı tarama"),
        (20, "P2", "SWE-bench Benchmark", "Gelişmiş AI", 5, "Çok Yüksek", "20-30 gün", "Tüm özellikler", "Endüstri standardı benchmark'ta rekabetçi skor"),
        (21, "P3", "Plan-to-Cloud Handoff", "API & Bulut", 6, "Yüksek", "7-10 gün", "Cloud Agent", "Yerel planlamayı buluta devretme"),
        (22, "P3", "Jira/Linear Entegrasyonu", "Entegrasyonlar", 5, "Orta", "5-7 gün", "Yok", "Proje yönetim araçlarıyla entegrasyon"),
        (23, "P3", "RTL Dil Desteği", "Uluslararasılaşma", 3, "Orta", "3-5 gün", "i18n Sistemi", "Sağdan sola yazılan diller için destek"),
        (24, "P3", "Masaüstü Uygulaması", "Geliştirici Deneyimi", 4, "Yüksek", "20-30 gün", "Electron/Tauri", "Goose tarzı native masaüstü uygulaması"),
    ]
    
    for i, item in enumerate(roadmap):
        row = 5 + i
        for col, val in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER if col <= 7 else ALIGN_LEFT
            cell.border = THIN_BORDER
            
            # Priority coloring
            if col == 2:
                if val == "P0":
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF842029")
                elif val == "P1":
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF664D03")
                elif val == "P2":
                    cell.fill = PatternFill(start_color="cfe2ff", end_color="cfe2ff", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF084298")
                else:
                    cell.fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FF495057")
    
    # Column widths
    widths = [5, 10, 32, 22, 12, 14, 14, 20, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    
    return ws


# ── Build Workbook ──
ws1 = wb.active
create_matrix_sheet(ws1, "Özellik Karşılaştırma Matrisi")
totals, max_score = create_score_sheet(wb.create_sheet())
create_gaps_sheet(wb.create_sheet())
create_skills_md_sheet(wb.create_sheet())
create_chart_sheet(wb.create_sheet(), totals, max_score)
create_priority_sheet(wb.create_sheet())

# Save
output_path = "/home/z/my-project/download/NeuroCLI_Karsilastirma_Matrisi_2026.xlsx"
wb.save(output_path)
print(f"✅ Dosya kaydedildi: {output_path}")

# Print summary
print("\n" + "="*60)
print("SKOR ÖZETİ")
print("="*60)
sorted_tools = sorted(TOOLS, key=lambda t: totals[t], reverse=True)
for rank, t in enumerate(sorted_tools, 1):
    pct = round((totals[t] / max_score) * 100, 1)
    bar = "█" * int(pct / 2) + "░" * (50 - int(pct / 2))
    print(f"  #{rank:2d} {t:20s} | {totals[t]:5.0f}/{max_score} | %{pct:5.1f} | {bar}")

print(f"\n  Toplam Özellik Sayısı: {len(features)}")
print(f"  Maksimum Skor: {max_score}")
print(f"  NeuroCLI Sıralaması: #{sorted_tools.index('NeuroCLI') + 1}")
